import openpyxl
import webbrowser
import tkinter

def openAndDisplay(word, filename):
    wrkbk = openpyxl.load_workbook(filename)  ##Part1
    sh = wrkbk.active
    i = word
    for j in range(1, sh.max_column):
        cell_obj = sh.cell(row=i, column=j)
        print( cell_obj.value)

# The question to get started
print("Welcome to ChemHelp. Please select the question you need and select the assigned number of the question")
question = int(input(" 1: Full details of the element, 2: Valance electron for a specific element, 3: want to learn about some topics"))
# To get the full data of a specific element

if question == 1:
    n = int(input("""Which of the following element would you like to have Atomic Number (amu, g/mol), Symbol,Name, Atomic weight :
Hydrogen	1	, Helium  	2	, Lithium 	3	, Beryllium	4	, Boron   	5	, Carbon  	6	, Nitrogen	7	, Oxygen  	8,	,
Fluorine	9	, Neon    	10	, Sodium  	11	, Magnesium	12	, Aluminum	13	, Silicon  	14	, Phosphorus	15	, Sulfur  	16,
Chlorine	17	, Potassium	18	, Argon   	19	, Calcium 	20	, Scandium	21	, Titanium	22	, Vanadium	23	, Chromium	24,
Manganese	25	, Iron    	26	, Nickel  	27	, Cobalt  	28	, Copper  	29	, Zinc    	30	, Gallium 	31	, Germanium	32
Arseni c	33	, Selenium	34	, Bromine 	35	, Krypton 	36	, Rubidium	37	, Strontium	38	, Yttrium 	39	, Zirconium	40
Niobium 	41	, Molybdenum	42	, Technetium	43	, Ruthenium	44	, Rhodium 	45	, Palladium	46	, Silver  	47      , Cadmium  	48	,
Indium   	49	, Tin     	50	, Antimony	51	, Iodine   	52	, Tellurium	53	, Xenon   	54	, Cesium  	55	, Barium  	56	,
Lanthanum	57	, Cerium  	58      , Praseodymium	59	, Neodymium	60	, Promethium	61	, Samarium	62	, Europium	63      , Gadolinium	64      ,
Terbium 	65	, Dysprosium	66	, Holmium 	67	, Erbium  	68	, Thulium 	69	, Ytterbium	70	, Lutetium	71	, Hafnium 	72	,
Tantalum	73	, Tungsten	74	, Rhenium 	75	, Osmium  	76	, Iridium   	77	, Platinum	78	, Gold    	79	, Mercury 	80	,
Thallium	81	, Lead    	82	, Bismuth 	83	, Polonium	84	, Astatine	85	, Radon   	86	, Francium	87	, Radium  	88	,
Actinium	89	, Protactinium	90	, Thorium 	91	, Neptunium	92	, Uranium 	93	, Plutonium	94	, Americium	95	, Berkelium	96	,
Curium  	97	, Nobelium	98	, Californium	99	, Einsteinium	100	, Hassium 	101	, Meitnerium	102	, Fermium 	103	, Mendelevium	104	,
Lawrencium	105	, Rutherfordium	106	, Bohrium 	107	, Dubnium 	108	, Seaborgium	109	, Ununnilium	110	, Unununium	111	, Ununbiium	112	,
——       	113	, Ununquadium	114	,

The number of the element you want to know about:-"""))
    print(openAndDisplay(n, "chemical element chart.xlsx"))

elif question == 2:
    n = int(input("""Which of the following elements would you get the valance electron of  :
Hydrogen	1	, Helium  	2	, Lithium 	3	, Beryllium	4	, Boron   	5	, Carbon  	6	, Nitrogen	7	, Oxygen  	8,	,
Fluorine	9	, Neon    	10	, Sodium  	11	, Magnesium	12	, Aluminum	13	, Silicon  	14	, Phosphorus	15	, Sulfur  	16,
Chlorine	17	, Potassium	18	, Argon   	19	, Calcium 	20	, Scandium	21	, Titanium	22	, Vanadium	23	, Chromium	24,
Manganese	25	, Iron    	26	, Nickel  	27	, Cobalt  	28	, Copper  	29	, Zinc    	30	, Gallium 	31	, Germanium	32
Arseni c	33	, Selenium	34	, Bromine 	35	, Krypton 	36	, Rubidium	37	, Strontium	38	, Yttrium 	39	, Zirconium	40
Niobium 	41	, Molybdenum	42	, Technetium	43	, Ruthenium	44	, Rhodium 	45	, Palladium	46	, Silver  	47      , Cadmium  	48	,
Indium   	49	, Tin     	50	, Antimony	51	, Iodine   	52	, Tellurium	53	, Xenon   	54	, Cesium  	55	, Barium  	56	,
Lanthanum	57	, Cerium  	58      , Praseodymium	59	, Neodymium	60	, Promethium	61	, Samarium	62	, Europium	63      , Gadolinium	64      ,
Terbium 	65	, Dysprosium	66	, Holmium 	67	, Erbium  	68	, Thulium 	69	, Ytterbium	70	, Lutetium	71	, Hafnium 	72	,
Tantalum	73	, Tungsten	74	, Rhenium 	75	, Osmium  	76	, Iridium   	77	, Platinum	78	, Gold    	79	, Mercury 	80	,
Thallium	81	, Lead    	82	, Bismuth 	83	, Polonium	84	, Astatine	85	, Radon   	86	, Francium	87	, Radium  	88	,
Actinium	89	, Protactinium	90	, Thorium 	91	, Neptunium	92	, Uranium 	93	, Plutonium	94	, Americium	95	, Berkelium	96	,
Curium  	97	, Nobelium	98	, Californium	99	, Einsteinium	100	, Hassium 	101	, Meitnerium	102	, Fermium 	103	, Mendelevium	104	,
Lawrencium	105	, Rutherfordium	106	, Bohrium 	107	, Dubnium 	108	, Seaborgium	109	, Ununnilium	110	, Unununium	111	, Ununbiium	112	,
    ——       	113	, Ununquadium	114	,

    The number of the element you want to know about:-"""))
    print(openAndDisplay(n, "Valance.xlsx"))

elif question == 3:
    print(""" The topic available are:
1 : electron configuration
2 : Periodic Table
3 : Different type of elements in periodic table """)
    which_topic_no = int(input("Which topic would you like to select"))

    if which_topic_no == 1: # 1 A electron Configuration
        print('''There are many explanations for the term "Electron Configuration". But the best way to describe it is by how wikipedia talks about it.
        " Electronic configurations describe each electron as moving independently in an orbital, in an average field created by all other orbitals."''')
# part A video
        askec = str(input("Would you like to see a video that explains the electron configuration"))
        askecl = askec.lower()
# part AB
        if askec == "y" or "yes":  # To watch a video explaining electron configuration
            webbrowser.open_new("https://www.youtube.com/watch?v=LyFzELsUgBU")

            askec2 = str(input("Would you like to find the electron configuaration of an element of you choice"))
            askec2l = askec2.lower()
# part  AC
            if askec2l == "y" or "yes": # to check
                n = int(input("""what element:
Hydrogen	1	, Helium  	2	, Lithium  	3	, Beryllium	4	, Boron   	5	, Carbon  	6	, Nitrogen	7	, Oxygen  	8	,
Fluorine	9	, Neon    	10	, Sodium   	11	, Magnesium	12	, Aluminum	13	, Silicon  	14	, Phosphorus	15	, Sulfur  	16	,
Chlorine	17	, Argon   	18	, Potassium	19	, Calcium  	20	, Scandium	21	, Titanium	22	, Vanadium	23	, Chromium	24	,
Manganese	25	, Iron    	26	, Cobalt  	27	, Nickel  	28	, Copper  	29	, Zinc    	30	, Gallium 	31	, Germanium	32	,
Arsenic 	33	, Selenium	34	, Bromine 	35	, Krypton 	36	, Rubidium	37	, Strontium	38	, Yttrium 	39	, Zirconium	40	,
Niobium  	41	, Molybdenum	42	, Technetium	43	, Ruthenium	44	, Rhodium  	45	, Palladium	46	, Silver   	47	, Cadmium  	48	,
Indium  	49	, Tin     	50	, Antimony	51	, Tellurium	52	, Iodine  	53	, Xenon   	54	, Cesium  	55	, Barium  	56	,
Lanthanum	57	, Cerium  	58	, Praseodymium	59	, Neodymium	60	, Promethium	61	, Samarium	62	, Europium	63	, Gadolinium	64	,
Terbium 	65	, Dysprosium	66	, Holmium 	67	, Erbium  	68	, Thulium 	69	, Ytterbium	70	, Lutetium	71	, Hafnium 	72	,
Tantalum	73	, Tungsten	74	, Rhenium  	75	, Osmium  	76	, Iridium 	77	, Platinum	78	, Gold    	79	, Mercury 	80	,
Thallium	81	, Lead    	82	, Bismuth 	83	, Polonium	84	, Astatine	85	, Radon   	86	, Francium	87	, Radium  	88	,
Actinium	89	, Thorium 	90	, Protactinium	91	, Uranium 	92	, Neptunium	93	, Plutonium	94	, Americium	95	, Curium   	96	,
Berkelium	97	, Californium	98	, Einsteinium	99	, Fermium 	100	, Mendelevium	101	, Nobelium	102	, Lawrencium	103	, Rutherfordium	104	,
Dubnium 	105	, Seaborgium	106	, Bohrium 	107	, Hassium  	108	, Meitnerium	109	, Darmstadtium	110	, Roentgenium	111	, Copernium	112	,
Nihonium	113	, Flerovium	114	, Moscovium	115	, Livermorium	116	, Tennessine	117	, Oganesson	118	,
                The number of the element you want to know about"""))

                print(openAndDisplay(n, "electron Configuration.xlsx"))

# part  B else
            else:  # no to search element
                print("Invalid")
                exit()
        else: # no to watch video
            exit()
            askec4 =str(input("Would you like to find the elctron configuaration of an element of you choice"))
            askec4l = askec4.lower()
            if askec4l == "y" or "yes": # to check
                n = int(input("""what element:
Hydrogen	1	, Helium  	2	, Lithium  	3	, Beryllium	4	, Boron   	5	, Carbon  	6	, Nitrogen	7	, Oxygen  	8	,
Fluorine	9	, Neon    	10	, Sodium   	11	, Magnesium	12	, Aluminum	13	, Silicon  	14	, Phosphorus	15	, Sulfur  	16	,
Chlorine	17	, Argon   	18	, Potassium	19	, Calcium  	20	, Scandium	21	, Titanium	22	, Vanadium	23	, Chromium	24	,
Manganese	25	, Iron    	26	, Cobalt  	27	, Nickel  	28	, Copper  	29	, Zinc    	30	, Gallium 	31	, Germanium	32	,
Arsenic 	33	, Selenium	34	, Bromine 	35	, Krypton 	36	, Rubidium	37	, Strontium	38	, Yttrium 	39	, Zirconium	40	,
Niobium  	41	, Molybdenum	42	, Technetium	43	, Ruthenium	44	, Rhodium  	45	, Palladium	46	, Silver   	47	, Cadmium  	48	,
Indium  	49	, Tin     	50	, Antimony	51	, Tellurium	52	, Iodine  	53	, Xenon   	54	, Cesium  	55	, Barium  	56	,
Lanthanum	57	, Cerium  	58	, Praseodymium	59	, Neodymium	60	, Promethium	61	, Samarium	62	, Europium	63	, Gadolinium	64	,
Terbium 	65	, Dysprosium	66	, Holmium 	67	, Erbium  	68	, Thulium 	69	, Ytterbium	70	, Lutetium	71	, Hafnium 	72	,
Tantalum	73	, Tungsten	74	, Rhenium  	75	, Osmium  	76	, Iridium 	77	, Platinum	78	, Gold    	79	, Mercury 	80	,
Thallium	81	, Lead    	82	, Bismuth 	83	, Polonium	84	, Astatine	85	, Radon   	86	, Francium	87	, Radium  	88	,
Actinium	89	, Thorium 	90	, Protactinium	91	, Uranium 	92	, Neptunium	93	, Plutonium	94	, Americium	95	, Curium   	96	,
Berkelium	97	, Californium	98	, Einsteinium	99	, Fermium 	100	, Mendelevium	101	, Nobelium	102	, Lawrencium	103	, Rutherfordium	104	,
Dubnium 	105	, Seaborgium	106	, Bohrium 	107	, Hassium  	108	, Meitnerium	109	, Darmstadtium	110	, Roentgenium	111	, Copernium	112	,
Nihonium	113	, Flerovium	114	, Moscovium	115	, Livermorium	116	, Tennessine	117	, Oganesson	118	,
                The number of the element you want to know about"""))
                print(openAndDisplay(n, "electron Configuration.xlsx"))

            else:
                print("invalid")
                exit()
    elif which_topic_no == 2:  # 2 Periodic Table
        print("""The periodic table, also known as the periodic table of the (chemical) elements, is a rows and columns arrangement of the chemical elements. It is widely used in chemistry, physics, and other sciences, and is generally seen as an icon of chemistry.
It is a graphic formulation of the periodic law, which states that the properties of the chemical elements exhibit an approximate periodic dependence on their atomic numbers. """)
        a = str(input("Do you want to see the Periodic table"))
        al = a.lower()
        if al == "yes" or "y":
            #map_bg_img = tkinter.PhotoImage(file="iPeriodic_Table_of_Elements_w_Chemical_Group_Block_PubChem.png")
            #print(map_bg_img)
            webbrowser.open_new(
                "https://www.google.com/search?q=periodic+table&rlz=1C1CHBF_enUS980US980&oq=pe&aqs=chrome.2.69i59l3j69i57j69i60l4.3554j0j7&sourceid=chrome&ie=UTF-8")
        else:
            print("Start again")
    elif which_topic_no == 3: # 3 A
        print("""On the periodic table, there are families which are groups of elements with similar properties.
These families are alkali metals, alkaline earth metals, transition metals, post-transition metals, metalloids, halogens, noble metals, and noble gases""")
        q = str(input("Do you want to see the different elements in Periodic table"))
        qt = q.lower()
        if qt == "yes" or "y":
            print("""
        The Alkali Metals (Group 1)
The alkali metals consist of all of the elements in group one with the exception of hydrogen. These elements are extremely reactive and for this reason, are usually found in compounds. In addition, they are water-sensitive (they react violently with water), so they must be stored in oil. The most reactive alkali metal is francium and it decreases as you go up the group. This means lithium is the least reactive. Physically, the alkali metal family is silvery, white, and light. They also have low melting and low boiling points.

        The Alkaline Earth Metals (Group 2)
The alkaline earth metals are the second most reactive family on the periodic table (following behind the alkali metals). Moreover, they are strong reducing agents which means they donate electrons in chemical reactions. They are also good thermal and electrical conductors. Physically, they have low density, low melting point, and a low boiling point.

        Rare Earth Metals: Lanthanides
Lanthanides are a family of rare earth metals that contain one valence electron in the 5d shell. They are highly reactive and a strong reducing agent in reactions. Furthermore, they are a silvery-bright metal and are relatively soft. They also have both high melting points and high boiling points. The rare earths include elements like neodymium and erbium.

        Rare Earth Metals: Actinides
Actinides are another family of rare earth metals. Like the lanthanides, these elements are highly reactive. They also have high electropositivity and are radioactive. Additionally, these elements contain paramagnetic, pyromorphic, and allotropic properties. Physically, they are very similar to lanthanides. They are silvery metals that are soft, malleable, and ductile.

        The Transition Metals (Groups 3-11)
The transition metals typically form two or more oxidation states. They have low ionization energies and high conductivity. In addition, they have high melting points, high boiling points, and high conductivity. Physically they are both metallic and malleable.

        Post Transition Metal
The post transition metals are located in between the transition metals and the metalloids. At standard temperature, they are in a solid state of matter. They tend to have a high density as well as high conductivity. Physically they are malleable and ductile.

        The Metalloids
The metalloids display properties of both metals and non-metals. For example, metals are good conductors and non-metals are poor conductors. This means metalloids are semiconductors (only conducts electricity at high temperatures.). Also, they are more brittle than metals but less brittle than non-metals. Physically they can be either shiny or dull and are typically ductile and malleable.

        The Halogens (Group 17)
The name halogen means “salt formers” in greek. This is evident in nature as halogens interact with metals to form various salts. On another note, the halogens are a unique group of elements. They are the only periodic family that contains elements in the three states of matter at standard temperature. There are 6 halogens and they are located in group 17. These elements include fluorine (F), chlorine (Cl),  bromine (Br), iodine (I), and astatine (At). They are highly reactive, highly electronegative, and highly toxic non-metals.

        Noble Metals
The noble metals consist of ruthenium (Ru), osmium (Os), rhodium (Rh), iridium (Ir), Pd, platinum (Pt), gold (Au), silver (Ag). Like the noble gases, they are inert due to having a complete valence shell. In addition, noble metals have catalytic tendencies. Also, they are very resistant to corrosion, tarnishing, and oxidation. Finally, like many of the other metals, they are soft and ductile.

        Noble Gases (Group 18)
The noble gases, also called aerogens, are inert gases. Some examples include argon, krypton, and neon. They can be found in group eighteen on the periodic table. Likewise, this means they have a complete valence shell. For this reason, they are stable and relatively unreactive. Furthermore, the noble gases have low boiling points and low melting points. Physically they are colorless and have no smell.""")
        else:
            print("Invalid")
